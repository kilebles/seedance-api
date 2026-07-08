from __future__ import annotations

import io
import uuid
from pathlib import Path

import openpyxl
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form, status
from loguru import logger
from sqlalchemy.ext.asyncio import AsyncSession

from src.core.database import get_db, AsyncSessionLocal
from src.core.settings import settings
from src.repositories import generation as generation_repo
from src.schemas.content import ContentItem, TextContent
from src.schemas.generation import AspectRatio, GenerationRequest, Resolution, TaskDB

router = APIRouter(prefix="/generations", tags=["generations"])


async def _get_admin_user_id(db: AsyncSession) -> uuid.UUID:
    from src.repositories import user as user_repo
    user = await user_repo.get_by_username(db, settings.admin_username)
    if user is None:
        raise HTTPException(status_code=status.HTTP_503_SERVICE_UNAVAILABLE, detail="Admin user not initialized")
    return user.id


@router.post(
    "/batch",
    response_model=list[TaskDB],
    status_code=status.HTTP_202_ACCEPTED,
    summary="Submit a batch of t2v generation tasks from an xlsx file",
)
async def create_batch(
    file: UploadFile = File(..., description="xlsx with columns: number, prompt"),
    model: str = Form(default="dreamina-seedance-2-0-fast-260128"),
    ratio: AspectRatio = Form(default=AspectRatio.ratio_16_9),
    resolution: Resolution = Form(default=Resolution.p720),
    duration: int = Form(default=8, ge=4, le=15),
    generate_audio: bool = Form(default=True),
    upscale_resolution: str | None = Form(default=None),
    db: AsyncSession = Depends(get_db),
) -> list[TaskDB]:
    if not file.filename or not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Only .xlsx files are supported")

    stem = Path(file.filename).stem
    if upscale_resolution:
        dir_name = f"SeeDance_{stem}_{resolution.value}_to_{upscale_resolution}"
    else:
        dir_name = f"SeeDance_{stem}_{resolution.value}"
    output_dir = f"output/batch/{dir_name}"
    logger.info("Batch upload | file={f} dir={d}", f=file.filename, d=dir_name)

    content = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content))
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Failed to parse xlsx: {exc}") from exc

    ws = wb.active
    headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
    if "prompt" not in headers:
        raise HTTPException(status_code=400, detail="xlsx must have a 'prompt' column")
    prompt_col = headers.index("prompt")

    raw_rows = list(ws.iter_rows(min_row=2, values_only=True))
    rows = [
        (str(r[0]).strip(), str(r[prompt_col]).strip())
        for r in raw_rows
        if r[0] and r[prompt_col]
    ]

    if not rows:
        raise HTTPException(status_code=400, detail="No data rows found in xlsx")

    logger.info("Batch: {n} rows parsed", n=len(rows))

    user_id = await _get_admin_user_id(db)

    batch_id = str(uuid.uuid4())
    async with AsyncSessionLocal() as session:
        from sqlalchemy import select, func as sqlfunc
        from src.models.generation import GenerationTask as GT
        result = await session.execute(select(sqlfunc.max(GT.batch_order)))
        max_order = result.scalar_one_or_none() or 0
    batch_order = max_order + 1

    logger.info("Batch: batch_id={bid} batch_order={bo}", bid=batch_id[:8], bo=batch_order)

    created_tasks: list[TaskDB] = []
    for number, prompt in rows:
        local_path = f"{output_dir}/{number}.mp4"
        content_items: list[ContentItem] = [TextContent(type="text", text=prompt)]

        request = GenerationRequest(
            model=model,
            content=content_items,
            ratio=ratio,
            resolution=resolution,
            duration=duration,
            generate_audio=generate_audio,
            upscale_resolution=upscale_resolution,
        )

        task = await generation_repo.create(
            db, user_id=user_id, request=request,
            name=number, local_path=local_path,
            batch_id=batch_id, batch_order=batch_order,
        )
        logger.debug("Batch: queued | name={n} id={id}", n=number, id=task.id)
        created_tasks.append(TaskDB.model_validate(task))

    logger.info("Batch: queued={q} batch_order={bo}", q=len(created_tasks), bo=batch_order)
    return created_tasks


@router.post("/batch/pause", summary="Pause a batch")
async def pause_batch(output_dir: str, db: AsyncSession = Depends(get_db)) -> dict:
    count = await generation_repo.set_batch_status(db, output_dir, from_statuses=["queued"], to_status="paused")
    logger.info("Batch paused | dir={d} tasks={n}", d=output_dir, n=count)
    return {"paused": count, "output_dir": output_dir}


@router.post("/batch/resume", summary="Resume a paused batch")
async def resume_batch(output_dir: str, db: AsyncSession = Depends(get_db)) -> dict:
    count = await generation_repo.set_batch_status(db, output_dir, from_statuses=["paused"], to_status="queued")
    logger.info("Batch resumed | dir={d} tasks={n}", d=output_dir, n=count)
    return {"resumed": count, "output_dir": output_dir}


@router.post("/batch/cancel", summary="Cancel a batch")
async def cancel_batch(output_dir: str, db: AsyncSession = Depends(get_db)) -> dict:
    count = await generation_repo.set_batch_status(db, output_dir, from_statuses=["queued", "paused"], to_status="failed")
    logger.info("Batch cancelled | dir={d} tasks={n}", d=output_dir, n=count)
    return {"cancelled": count, "output_dir": output_dir}


@router.post("/batch/retry-failed", summary="Re-queue failed tasks in a batch")
async def retry_failed(output_dir: str, db: AsyncSession = Depends(get_db)) -> dict:
    count = await generation_repo.set_batch_status(db, output_dir, from_statuses=["failed"], to_status="queued")
    logger.info("Batch retry | dir={d} tasks={n}", d=output_dir, n=count)
    return {"retried": count, "output_dir": output_dir}
